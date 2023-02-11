import requests
import openpyxl
from datetime import datetime

# Define the Binance API endpoint
symbol = "BTCUSDT"
interval = "1d"
url = f"https://api.binance.com/api/v1/klines?symbol={symbol}&interval={interval}"

# Make a GET request to the API endpoint
response = requests.get(url)

# Parse the response JSON data
data = response.json()

# Extract the relevant information from the response data
coin = "BTC"
market_cap = data["quoteVolume"]
last_traded_price = data["lastPrice"]
change = data["priceChangePercent"]

# Write the information to a new spreadsheet
wb = openpyxl.Workbook()
sheet = wb.active

# Define the column headers
sheet.cell(row=1, column=1, value="Coin")
sheet.cell(row=1, column=2, value="24h change")
sheet.cell(row=1, column=3, value="Market cap")
sheet.cell(row=1, column=4, value="Last traded price")

# Write the data to the next row
sheet.cell(row=2, column=1, value=coin)
sheet.cell(row=2, column=2, value=change)
sheet.cell(row=2, column=3, value=market_cap)
sheet.cell(row=2, column=4, value=last_traded_price)

# Generate the file name as today's date in the format yyyymmdd
today = datetime.now().strftime("%Y%m%d")
filename = f"binance_data_{coin}{today}.xlsx"

# Save the spreadsheet to disk
wb.save(filename)